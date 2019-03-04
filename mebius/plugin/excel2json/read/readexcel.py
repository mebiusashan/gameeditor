# coding=UTF-8

from ..data.vo import ExcelMeta
from ..data.vo import MyExcel
from ..data.vo import SheetData

from ..error import openExcelErr

from openpyxl import load_workbook


def open(path):
    '''
    打开excel文件
    :param path:excel文件绝对路径
    :return:解析结果
    '''
    wb = load_workbook(path)
    rel = __inspectExcel(wb)
    m = ExcelMeta
    if len(rel) != 0:  # 存在错误
        m.hasErr = True;
        m.err = rel;
        return m;
    excel = __readExcel(wb)
    m.myExcel = excel
    m.wb = wb
    return m


def __readExcel(wb):
    m = MyExcel();
    sheets = wb.sheetnames;
    for i in sheets:
        sh = wb.get_sheet_by_name(i)
        if i == 'config':
            m.defConfig.addTypeLen(sh['B1'].value, sh['B2'].value)
        elif i == 'id_type':
            l = sh.max_row + 1
            for t in range(2, l):
                m.addIDType(sh['A' + str(t)].value, sh['B' + str(t)].value, sh['C' + str(t)].value)
        else:
            sd = __createSheetData(sh)
            m.addSheet(i, sd)
    return m


def __createSheetData(sh):
    sd = SheetData()
    sd.setSheetConfig(sh['B1'].value, sh.title, sh['D1'].value)
    # 处理表结构 SheetStruct
    for column in sh.columns:
        if column[1].value == None or column[2].value == None or column[3].value == None or column[4].value == None or \
                column[5].value == None:
            break
        sd.addSheetStruct(column[1].value, column[2].value, column[3].value, column[4].value, "",
                          column[5].value)
    # 处理中的数据
    keys = sh["2"]  # 第二行，表示数据的key
    row = sh.max_row + 1
    for t in range(7, row):
        data = sh[t]
        d = {}
        colnum = len(data)
        for p in range(0, colnum):
            if keys[p].value != None:
                d[keys[p].value] = data[p].value
        sd.addData(d)
    return sd


def __inspectExcel(workbook):
    hasConfigSheet = False
    hasIDTypeSheet = False
    sheets = workbook.sheetnames;
    for i in sheets:
        if i == 'config':
            hasConfigSheet = True;
        elif i == 'id_type':
            hasIDTypeSheet = True;

    rel = []
    if hasConfigSheet == False:
        rel.append(openExcelErr['NotFoundConfigSheet'])
    if hasIDTypeSheet == False:
        rel.append(openExcelErr['NotFoundIDTypeSheet'])
    return rel
